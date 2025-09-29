# core/views.py
import io, base64, csv, re, tempfile
from decimal import Decimal
from datetime import datetime
from collections import Counter


# Django
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q, DecimalField, Value as V
from django.db.models.functions import Coalesce, ExtractMonth
from django.template.loader import get_template, render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.html import format_html
from django.views.decorators.http import require_http_methods
from django.forms import inlineformset_factory
from .models import UniBookingCard, HotelBooking, Room # تأكد من استيراد Room
from .forms import HotelBookingForm, RoomForm # تأكد من استيراد RoomForm


# Models & Forms
from .models import (
    UniBookingCard, HotelBooking, Payment,
    FlightBooking, TransferBooking, VisaBooking
)
from .forms import (
    UniBookingCardForm, HotelBookingForm, PaymentForm,
    FlightBookingForm, 
    TransferBookingForm, VisaBookingForm,
    VoucherUploadForm, VoucherConfirmForm
)



# External
from xhtml2pdf import pisa
import qrcode
from qrcode.image.pil import PilImage
import pdfplumber
from docx import Document

# ===================== Helpers =====================
def _build_qr_data_url(text: str) -> str:
    qr = qrcode.QRCode(version=1, box_size=4, border=1)
    qr.add_data(text); qr.make(fit=True)
    img: PilImage = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO(); img.save(buf, format='PNG')
    return f"data:image/png;base64,{base64.b64encode(buf.getvalue()).decode('ascii')}"

def _render_pdf_from_template(template_name: str, context: dict) -> bytes:
    html = get_template(template_name).render(context)
    out = io.BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=out, encoding='utf-8')
    return None if pisa_status.err else out.getvalue()

def _cards_base_qs(request):
    return UniBookingCard.objects.all() if request.user.is_superuser else UniBookingCard.objects.filter(created_by=request.user)

# ===================== Dashboard / Cards =====================
@login_required
def dashboard(request):
    qs = _cards_base_qs(request)
    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(customer_name__icontains=q) |
            Q(ub_code__icontains=q) |
            Q(mobile__icontains=q)
        )
    from django.core.paginator import Paginator
    paginator = Paginator(qs.order_by("-created_at"), 12)
    page_obj = paginator.get_page(request.GET.get("page"))
    return render(request, "core/dashboard.html", {"cards": page_obj, "q": q})

@login_required
def card_create(request):
    if request.method == 'POST':
        form = UniBookingCardForm(request.POST)
        if form.is_valid():
            card = form.save(commit=False)
            card.created_by = request.user
            card.save()
            messages.success(request, "تم إنشاء كارت العميل.")
            return redirect('card_detail', pk=card.pk)
        else:
            errs = [f"{f}: {', '.join(e)}" for f, e in form.errors.items()]
            messages.error(request, format_html("تحقق من المدخلات: {}", "; ".join(errs)))
    else:
        form = UniBookingCardForm()
    return render(request, 'core/card_form.html', {'form': form})
@login_required
def card_detail(request, pk):
    card = get_object_or_404(UniBookingCard, pk=pk)

    # حجوزات
    hotel_bookings = card.hotelbooking_set.all()
    flight_bookings = card.flightbooking_set.all()
    transfer_bookings = card.transferbooking_set.all()
    visa_bookings = card.visabooking_set.all()

    # إجماليات
    totals = {
        "total_sell": card.total_sell,
        "total_net": card.total_net,
        "total_paid": card.total_paid,
        "total_remaining": card.total_remaining,
        "total_profit": card.total_profit,
    }

    return render(request, "core/card_detail.html", {
        "card": card,
        "hotel_bookings": hotel_bookings,
        "flight_bookings": flight_bookings,
        "transfer_bookings": transfer_bookings,
        "visa_bookings": visa_bookings,
        "totals": totals,
    })


    # ===================== Cards Bulk Export =====================
@login_required
def cards_bulk_export(request):
    ids = request.GET.getlist('ids[]') or request.GET.getlist('ids')
    qs = _cards_base_qs(request)
    if ids:
        qs = qs.filter(id__in=ids)
    qs = qs.annotate(
        hotels_count=Count('hotelbooking', distinct=True),
        flights_count=Count('flightbooking', distinct=True),
        transfers_count=Count('transferbooking', distinct=True),
        visas_count=Count('visabooking', distinct=True),
    )

    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = 'attachment; filename="cards_export.csv"'
    writer = csv.writer(resp)
    writer.writerow([
        'ID','UB Code','Customer','Mobile','Nationality','Country',
        'Hotels','Flights','Transfers','Visas','Created At'
    ])
    for c in qs:
        writer.writerow([
            c.id, c.ub_code, c.customer_name, c.mobile or '',
            c.nationality or '', c.country or '',
            c.hotels_count, c.flights_count, c.transfers_count, c.visas_count,
            c.created_at.isoformat() if c.created_at else ''
        ])
    return resp
# ===================== Cards Bulk Delete =====================
@login_required
def cards_bulk_delete(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Invalid method'}, status=405)

    ids = request.POST.getlist('ids[]') or request.POST.getlist('ids')
    if not ids:
        return JsonResponse({'ok': False, 'error': 'No IDs'}, status=400)

    qs = _cards_base_qs(request).filter(id__in=ids)
    count = qs.count()
    if not count:
        return JsonResponse({'ok': False, 'error': 'No cards found or not permitted'}, status=404)

    qs.delete()
    return JsonResponse({'ok': True, 'deleted': count})

# ===================== Hotel CRUD =====================

@login_required
def hotel_create(request, card_pk):
    card = get_object_or_404(UniBookingCard, pk=card_pk)

    if not request.user.is_superuser and card.created_by != request.user:
        messages.error(request, "غير مسموح.")
        return redirect('dashboard')

    # إنشاء Formset للغرف المرتبطة بالحجز
    # extra=1 يعني أنه سيظهر فورم غرفة واحدة بشكل افتراضي
    RoomFormSet = inlineformset_factory(
        HotelBooking, 
        Room, 
        form=RoomForm, 
        extra=1, 
        can_delete=False
    )

    if request.method == 'POST':
        form = HotelBookingForm(request.POST)
        room_formset = RoomFormSet(request.POST, prefix='rooms')

        if form.is_valid() and room_formset.is_valid():
            hb = form.save(commit=False)
            hb.card = card
            hb.employee_name = request.user.get_full_name() or request.user.username

            if hb.checkin and hb.checkout and hb.checkout > hb.checkin:
                hb.nights = (hb.checkout - hb.checkin).days
            
            # يجب حفظ الحجز أولاً قبل حفظ الغرف المرتبطة به
            hb.save()

            # ربط الفورمست بالحجز الذي تم حفظه ثم حفظ الغرف
            room_formset.instance = hb
            room_formset.save()

            messages.success(request, f"تم حفظ حجز الفندق. فوّتشر: {hb.voucher_code}")
            return redirect('hotel_payment', booking_pk=hb.pk)
        else:
            # عرض رسائل الأخطاء في حالة عدم الصحة
            if not form.is_valid():
                messages.error(request, "يرجى التحقق من بيانات الحجز الأساسية.")
            if not room_formset.is_valid():
                messages.error(request, "يرجى التحقق من بيانات أسماء النزلاء في الغرف.")

    else:
        form = HotelBookingForm()
        room_formset = RoomFormSet(prefix='rooms')

    return render(request, 'core/hotel_form.html', {
        'card': card, 
        'form': form,
        'room_formset': room_formset # إرسال الفورمست إلى القالب
    })


@login_required
def hotel_payment(request, booking_pk):
    booking = get_object_or_404(HotelBooking, pk=booking_pk)
    payments = booking.payments.order_by("-created_at")

    # حساب الإجماليات بناءً على الدفعات المسجلة
    total_paid = payments.aggregate(total=Sum("paid_amount"))["total"] or Decimal("0.00")
    
    # سعر البيع والنت يؤخذ من الحجز نفسه
    total_sell = booking.sell or Decimal("0.00")
    total_net = booking.net or Decimal("0.00")
    
    remaining = total_sell - total_paid
    profit = total_sell - total_net

    if request.method == "POST":
        # نمرر معلومات إضافية للفورم، مثل هل توجد دفعات سابقة أم لا
        form = PaymentForm(
            request.POST,
            request.FILES,
            booking=booking,
            has_existing_payments=payments.exists()
        )
        
        if form.is_valid():
            new_payment = form.save(commit=False)
            new_payment.booking_hotel = booking
            new_payment.employee_name = request.user.get_full_name() or request.user.username

            # إذا كانت هذه أول دفعة، قم بتحديث سعر الحجز الأساسي
            if not payments.exists():
                booking.net = form.cleaned_data.get('net_price', 0)
                booking.sell = form.cleaned_data.get('sell_price', 0)
                booking.save()
                
                remaining_after_sell_set = booking.sell - new_payment.paid_amount
                messages.success(
                    request,
                    f"تم تسجيل أول دفعة وتحديد سعر الحجز. المتبقي الآن: {remaining_after_sell_set:.2f} KWD"
                )
            else:
                remaining_after_payment = remaining - new_payment.paid_amount
                messages.success(
                    request,
                    f"تم تسجيل دفعة جديدة. المتبقي الآن: {remaining_after_payment:.2f} KWD"
                )

            new_payment.save()
            return redirect("hotel_payment", booking_pk=booking.pk)
        else:
            messages.error(request, "من فضلك صحح الأخطاء في النموذج.")
    else:
        form = PaymentForm(booking=booking, has_existing_payments=payments.exists())

    context = {
        "booking": booking,
        "form": form,
        "payments": payments,
        "remaining": remaining,
        "total_paid": total_paid,
        "total_sell": total_sell,
        "total_net": total_net,   # ✅ أضفت النت للـ context
        "profit": profit,         # ✅ الربح متحسب وموجود
    }
    return render(request, "core/hotel_payment.html", context)


@login_required
def hotel_voucher(request, booking_pk):
    booking = get_object_or_404(HotelBooking, pk=booking_pk)
    if not request.user.is_superuser and booking.card.created_by != request.user:
        messages.error(request, "غير مسموح.")
        return redirect('dashboard')
    voucher_url = request.build_absolute_uri(reverse('hotel_voucher', args=[booking.pk]))
    qr_data_url = _build_qr_data_url(voucher_url)
    return render(request, 'core/voucher.html', {
        'booking': booking, 'today': timezone.localdate(), 'nights': booking.nights, 'qr_data_url': qr_data_url,
    })


@login_required
def hotel_voucher_pdf(request, booking_pk):
    booking = get_object_or_404(HotelBooking, pk=booking_pk)
    if not request.user.is_superuser and booking.card.created_by != request.user:
        messages.error(request, "غير مسموح.")
        return redirect('dashboard')
    voucher_url = request.build_absolute_uri(reverse('hotel_voucher', args=[booking.pk]))
    qr_data_url = _build_qr_data_url(voucher_url)
    pdf_bytes = _render_pdf_from_template('core/voucher_pdf.html', {
        'booking': booking, 'today': timezone.localdate(), 'nights': booking.nights, 'qr_data_url': qr_data_url,
    })
    if pdf_bytes is None:
        return HttpResponse("PDF render error", status=500)
    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="Voucher_{booking.voucher_code}.pdf"'
    return resp
# ===================== Flight CRUD =====================

@login_required
def flight_create(request, card_pk):
    card = get_object_or_404(UniBookingCard, pk=card_pk)

    if request.method == "POST":
        form = FlightBookingForm(request.POST, request.FILES)
        if form.is_valid():
            booking = form.save(commit=False)
            booking.card = card

            # 🟢 توليد الكود الداخلي هنا
            if not booking.booking_code:
                today = timezone.now().strftime("%Y%m%d")
                emp_code = request.user.username.upper()[:5]  # كود الموظف
                booking.booking_code = f"F{today}{booking.pnr.upper()}{emp_code}"

            booking.save()
            return redirect("card_detail", pk=card.pk)
    else:
        form = FlightBookingForm()

    return render(request, "core/flight_form.html", {"form": form, "card": card})





@login_required
def flight_voucher(request, booking_pk):
    b = get_object_or_404(FlightBooking, pk=booking_pk)
    if not request.user.is_superuser and b.card.created_by != request.user:
        messages.error(request, "غير مسموح.")
        return redirect("dashboard")

    url = request.build_absolute_uri(reverse("flight_voucher", args=[b.pk]))
    qr = _build_qr_data_url(url)

    return render(request, "core/flight_voucher.html", {
        "b": b,
        "today": timezone.localdate(),
        "qr_data_url": qr,
    })


@login_required
def flight_voucher_pdf(request, booking_pk):
    b = get_object_or_404(FlightBooking, pk=booking_pk)
    if not request.user.is_superuser and b.card.created_by != request.user:
        messages.error(request, "غير مسموح.")
        return redirect('dashboard')

    url = request.build_absolute_uri(reverse('flight_voucher', args=[b.pk]))
    qr = _build_qr_data_url(url)

    pdf = _render_pdf_from_template('core/flight_voucher_pdf.html', {
        'b': b, 'today': timezone.localdate(), 'qr_data_url': qr
    })
    if pdf is None:
        return HttpResponse("PDF render error", status=500)

    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="Flight_{b.voucher_code}.pdf"'
    return resp

@login_required
def flight_payment(request, booking_pk):
    booking = get_object_or_404(FlightBooking, pk=booking_pk)

    context = {
        "booking": booking,
        "net": booking.net_price,
        "sell": booking.sell_price,
        "paid": booking.sell_price,     # 🟢 الطيران مدفوع بالكامل
        "remaining": Decimal("0.00"),   # 🟢 مفيش أقساط
        "profit": booking.profit,       # 🟢 يجيب من الـ property
    }
    return render(request, "core/flight_payment.html", context)




# ===================== Transfer CRUD =====================

@login_required
def transfer_create(request, card_pk):
    card = get_object_or_404(UniBookingCard, pk=card_pk)

    if not request.user.is_superuser and card.created_by != request.user:
        messages.error(request, "غير مسموح.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = TransferBookingForm(request.POST)
        if form.is_valid():
            b = form.save(commit=False)
            b.card = card
            b.employee_name = request.user.get_full_name() or request.user.username
            b.save()
            messages.success(request, f"تم حفظ حجز التوصيل. فوّتشر: {b.voucher_code}")
            return redirect('transfer_voucher', booking_pk=b.pk)
    else:
        form = TransferBookingForm()

    return render(request, 'core/transfer_form.html', {'card': card, 'form': form})


@login_required
def transfer_voucher(request, booking_pk):
    b = get_object_or_404(TransferBooking, pk=booking_pk)
    if not request.user.is_superuser and b.card.created_by != request.user:
        messages.error(request, "غير مسموح.")
        return redirect('dashboard')

    url = request.build_absolute_uri(reverse('transfer_voucher', args=[b.pk]))
    qr = _build_qr_data_url(url)

    return render(request, 'core/transfer_voucher.html', {
        'b': b, 'today': timezone.localdate(), 'qr_data_url': qr
    })


@login_required
def transfer_voucher_pdf(request, booking_pk):
    b = get_object_or_404(TransferBooking, pk=booking_pk)
    if not request.user.is_superuser and b.card.created_by != request.user:
        messages.error(request, "غير مسموح.")
        return redirect('dashboard')

    url = request.build_absolute_uri(reverse('transfer_voucher', args=[b.pk]))
    qr = _build_qr_data_url(url)

    pdf = _render_pdf_from_template('core/transfer_voucher_pdf.html', {
        'b': b, 'today': timezone.localdate(), 'qr_data_url': qr
    })
    if pdf is None:
        return HttpResponse("PDF render error", status=500)

    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="Transfer_{b.voucher_code}.pdf"'
    return resp
# ===================== Visa CRUD =====================

@login_required
def visa_create(request, card_pk):
    card = get_object_or_404(UniBookingCard, pk=card_pk)

    if not request.user.is_superuser and card.created_by != request.user:
        messages.error(request, "غير مسموح.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = VisaBookingForm(request.POST)
        if form.is_valid():
            b = form.save(commit=False)
            b.card = card
            b.employee_name = request.user.get_full_name() or request.user.username
            b.save()
            messages.success(request, f"تم حفظ حجز الفيزا. فوّتشر: {b.voucher_code}")
            return redirect('visa_voucher', booking_pk=b.pk)
    else:
        form = VisaBookingForm()

    return render(request, 'core/visa_form.html', {'card': card, 'form': form})


@login_required
def visa_voucher(request, booking_pk):
    b = get_object_or_404(VisaBooking, pk=booking_pk)
    if not request.user.is_superuser and b.card.created_by != request.user:
        messages.error(request, "غير مسموح.")
        return redirect('dashboard')

    url = request.build_absolute_uri(reverse('visa_voucher', args=[b.pk]))
    qr = _build_qr_data_url(url)

    return render(request, 'core/visa_voucher.html', {
        'b': b, 'today': timezone.localdate(), 'qr_data_url': qr
    })


@login_required
def visa_voucher_pdf(request, booking_pk):
    b = get_object_or_404(VisaBooking, pk=booking_pk)
    if not request.user.is_superuser and b.card.created_by != request.user:
        messages.error(request, "غير مسموح.")
        return redirect('dashboard')

    url = request.build_absolute_uri(reverse('visa_voucher', args=[b.pk]))
    qr = _build_qr_data_url(url)

    pdf = _render_pdf_from_template('core/visa_voucher_pdf.html', {
        'b': b, 'today': timezone.localdate(), 'qr_data_url': qr
    })
    if pdf is None:
        return HttpResponse("PDF render error", status=500)

    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="Visa_{b.voucher_code}.pdf"'
    return resp
# ===================== Reports + CSV =====================

@login_required
def reports(request):
    date_from = parse_date(request.GET.get('from') or '')
    date_to   = parse_date(request.GET.get('to') or '')
    employee  = (request.GET.get('employee') or '').strip()
    kind      = (request.GET.get('kind') or '').strip()

    base_filter = {} if request.user.is_superuser else {'card__created_by': request.user}
    data, labels = [], []

    def _append(qs, kname, formatter):
        if date_from: qs = qs.filter(created_at__date__gte=date_from)
        if date_to:   qs = qs.filter(created_at__date__lte=date_to)
        if employee:  qs = qs.filter(employee_name__icontains=employee)
        for r in qs.select_related('card')[:1000]:
            data.append({
                'kind': kname, 'ref': r.booking_ref, 'voucher': r.voucher_code,
                'employee': r.employee_name,
                'customer': r.card.customer_name if r.card_id else '',
                'created': r.created_at, 'extra': formatter(r)
            })
        labels.append((kname, qs.count()))

    if kind in ('hotel',''):    
        _append(HotelBooking.objects.filter(**base_filter), 'Hotel', lambda r: f'{r.hotel_name} / {r.country or "-"}')
    if kind in ('flight',''):   
        _append(FlightBooking.objects.filter(**base_filter), 'Flight', lambda r: f'{r.from_city}→{r.to_city} {r.depart_date}')
    if kind in ('transfer',''): 
        _append(TransferBooking.objects.filter(**base_filter), 'Transfer', lambda r: f'{r.pickup}→{r.dropoff} {r.date}')
    if kind in ('visa',''):     
        _append(VisaBooking.objects.filter(**base_filter), 'Visa', lambda r: f'{r.visa_type or "-"} / {r.nationality or "-"}')

    data.sort(key=lambda x: x['created'], reverse=True)
    chart_labels = [lbl for lbl,_ in labels]
    chart_values = [cnt for _,cnt in labels]

    return render(request, 'core/reports.html', {
        'rows': data, 'chart_labels': chart_labels, 'chart_values': chart_values,
        'q_from': request.GET.get('from',''), 'q_to': request.GET.get('to',''),
        'q_employee': employee, 'q_kind': kind,
    })


@login_required
def reports_export_csv(request):
    date_from = parse_date(request.GET.get('from') or '')
    date_to   = parse_date(request.GET.get('to') or '')
    employee  = (request.GET.get('employee') or '').strip()
    kind      = (request.GET.get('kind') or '').strip()
    base_filter = {} if request.user.is_superuser else {'card__created_by': request.user}
    rows = []

    def _append(qs, kname, formatter):
        if date_from: qs = qs.filter(created_at__date__gte=date_from)
        if date_to:   qs = qs.filter(created_at__date__lte=date_to)
        if employee:  qs = qs.filter(employee_name__icontains=employee)
        for r in qs.select_related('card')[:10000]:
            rows.append([kname, r.booking_ref, r.voucher_code, r.employee_name,
                         r.card.customer_name if r.card_id else '',
                         r.created_at.isoformat(), formatter(r)])

    if kind in ('hotel',''):    
        _append(HotelBooking.objects.filter(**base_filter),'Hotel',lambda r:f'{r.hotel_name}/{r.country}')
    if kind in ('flight',''):   
        _append(FlightBooking.objects.filter(**base_filter),'Flight',lambda r:f'{r.from_city}→{r.to_city} {r.depart_date}')
    if kind in ('transfer',''): 
        _append(TransferBooking.objects.filter(**base_filter),'Transfer',lambda r:f'{r.pickup}→{r.dropoff} {r.date}')
    if kind in ('visa',''):     
        _append(VisaBooking.objects.filter(**base_filter),'Visa',lambda r:f'{r.visa_type}/{r.nationality}')

    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = 'attachment; filename="reports.csv"'
    writer = csv.writer(resp)
    writer.writerow(['Type','Booking Ref','Voucher','Employee','Customer','Created At','Extra'])
    writer.writerows(rows)
    return resp


# ===================== Reports Export XLSX =====================

@login_required
def reports_export_xlsx(request):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    date_from = parse_date(request.GET.get('from') or '')
    date_to   = parse_date(request.GET.get('to') or '')
    employee  = (request.GET.get('employee') or '').strip()
    kind      = (request.GET.get('kind') or '').strip()
    base_filter = {} if request.user.is_superuser else {'card__created_by': request.user}
    rows = []

    def _append(qs, kname, formatter):
        if date_from: qs = qs.filter(created_at__date__gte=date_from)
        if date_to:   qs = qs.filter(created_at__date__lte=date_to)
        if employee:  qs = qs.filter(employee_name__icontains=employee)
        for r in qs.select_related('card')[:10000]:
            rows.append([kname,r.booking_ref,r.voucher_code,r.employee_name,
                         r.card.customer_name if r.card_id else '',
                         r.created_at.isoformat(),formatter(r)])

    if kind in ('hotel',''):    
        _append(HotelBooking.objects.filter(**base_filter),'Hotel',lambda r:f'{r.hotel_name}/{r.country}')
    if kind in ('flight',''):   
        _append(FlightBooking.objects.filter(**base_filter),'Flight',lambda r:f'{r.from_city}→{r.to_city} {r.depart_date}')
    if kind in ('transfer',''): 
        _append(TransferBooking.objects.filter(**base_filter),'Transfer',lambda r:f'{r.pickup}→{r.dropoff} {r.date}')
    if kind in ('visa',''):     
        _append(VisaBooking.objects.filter(**base_filter),'Visa',lambda r:f'{r.visa_type}/{r.nationality}')

    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"
    header = ['Type','Booking Ref','Voucher','Employee','Customer','Created At','Extra']
    ws.append(header)
    for row in rows: ws.append(row)
    for i,title in enumerate(header,1):
        ws.column_dimensions[get_column_letter(i)].width = max(12,len(title)+2)

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="reports.xlsx"'
    wb.save(resp)
    return resp


# ===================== Reports Export PDF =====================

@login_required
def reports_export_pdf(request):
    date_from = parse_date(request.GET.get('from') or '')
    date_to   = parse_date(request.GET.get('to') or '')
    employee  = (request.GET.get('employee') or '').strip()
    kind      = (request.GET.get('kind') or '').strip()
    base_filter = {} if request.user.is_superuser else {'card__created_by': request.user}
    rows = []

    def _append(qs,kname,formatter):
        if date_from: qs = qs.filter(created_at__date__gte=date_from)
        if date_to:   qs = qs.filter(created_at__date__lte=date_to)
        if employee:  qs = qs.filter(employee_name__icontains=employee)
        for r in qs.select_related('card')[:10000]:
            rows.append({'type':kname,'ref':r.booking_ref,'voucher':r.voucher_code,
                         'employee':r.employee_name,
                         'customer':r.card.customer_name if r.card_id else '',
                         'created':r.created_at,'extra':formatter(r)})

    if kind in ('hotel',''):    
        _append(HotelBooking.objects.filter(**base_filter),'Hotel',lambda r:f'{r.hotel_name}/{r.country}')
    if kind in ('flight',''):   
        _append(FlightBooking.objects.filter(**base_filter),'Flight',lambda r:f'{r.from_city}→{r.to_city} {r.depart_date}')
    if kind in ('transfer',''): 
        _append(TransferBooking.objects.filter(**base_filter),'Transfer',lambda r:f'{r.pickup}→{r.dropoff} {r.date}')
    if kind in ('visa',''):     
        _append(VisaBooking.objects.filter(**base_filter),'Visa',lambda r:f'{r.visa_type}/{r.nationality}')

    rows.sort(key=lambda x: x['created'], reverse=True)
    pdf_bytes = _render_pdf_from_template('core/reports_pdf.html',{
        'rows':rows,'q_from':request.GET.get('from',''),'q_to':request.GET.get('to',''),
        'q_employee':employee,'q_kind':kind,'generated_at':timezone.now(),'user':request.user,
    })
    if pdf_bytes is None:
        return HttpResponse("PDF render error",status=500)

    resp = HttpResponse(pdf_bytes,content_type='application/pdf')
    resp['Content-Disposition']='attachment; filename="reports.pdf"'
    return resp
# ===================== Voucher Upload / Parse =====================

DATE_PATTERNS = [
    r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
    r'([A-Za-z]{3}\s?\d{1,2},?\s?\d{4})',
]

def parse_dates(text):
    def _norm(s):
        for fmt in ("%d/%m/%Y","%d-%m-%Y","%m/%d/%Y","%m-%d-%Y","%b %d %Y","%b %d, %Y"):
            try:
                return datetime.strptime(s.replace(",",""), fmt).date()
            except:
                pass
        return None
    checkin = checkout = None
    for key in ["check in","check-in","arrival","from"]:
        m = re.search(key+r".{0,20}"+DATE_PATTERNS[0], text, re.IGNORECASE) or \
            re.search(key+r".{0,20}"+DATE_PATTERNS[1], text, re.IGNORECASE)
        if m: checkin = _norm(m.group(1)); break
    for key in ["check out","check-out","departure","to"]:
        m = re.search(key+r".{0,20}"+DATE_PATTERNS[0], text, re.IGNORECASE) or \
            re.search(key+r".{0,20}"+DATE_PATTERNS[1], text, re.IGNORECASE)
        if m: checkout = _norm(m.group(1)); break
    return checkin, checkout


def extract_text_from_pdf(fobj):
    with pdfplumber.open(fobj) as pdf:
        return "\n".join([p.extract_text() or "" for p in pdf.pages])


def extract_text_from_docx(fobj):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp:
        tmp.write(fobj.read())
        tmp.flush()
        doc = Document(tmp.name)
    return "\n".join([p.text for p in doc.paragraphs])


def smart_parse(text):
    data = {}
    def _grab(pattern, group=2):
        m = re.search(pattern,text,re.IGNORECASE)
        return m.group(group).strip() if m else None
    data["booking_ref"] = _grab(r'(booking\s*ref|reservation\s*id)\s*[:\-]?\s*([A-Z0-9\-]+)')
    data["voucher_code"]= _grab(r'(voucher\s*code|voucher)\s*[:\-]?\s*([A-Z0-9\-]+)')
    data["customer_display_name"]= _grab(r'(guest|holder|customer)\s*[:\-]?\s*([A-Za-z \-]+)')
    data["hotel_name"]= _grab(r'(hotel)\s*[:\-]?\s*([A-Za-z0-9 ,\-\.\(\)]+)')
    data["hotel_address"]= _grab(r'(address)\s*[:\-]?\s*(.+)',2)
    data["provider_name"]= _grab(r'(provider|supplier)\s*[:\-]?\s*([A-Za-z0-9 \-]+)')
    data["country"]= _grab(r'(country)\s*[:\-]?\s*([A-Za-z \-]+)')
    data["meal_plan"]= _grab(r'(meal|board)\s*[:\-]?\s*([A-Za-z \-]+)')
    data["room_type"]= _grab(r'(room\s*type)\s*[:\-]?\s*(.+)',2)

    m = re.search(r'(nights?)\s*[:\-]?\s*(\d+)',text,re.IGNORECASE)
    data["nights"]= int(m.group(2)) if m else None
    m = re.search(r'(rooms?)\s*[:\-]?\s*(\d+)',text,re.IGNORECASE)
    data["rooms_count"]= int(m.group(2)) if m else None
    ci,co = parse_dates(text)
    data["checkin"],data["checkout"]=ci,co
    return data


@require_http_methods(["GET","POST"])
def voucher_upload(request):
    if request.method=="POST":
        form = VoucherUploadForm(request.POST, request.FILES)
        if form.is_valid():
            f = form.cleaned_data["file"]
            ext = (f.name.split(".")[-1] or "").lower()
            try:
                text = extract_text_from_pdf(f) if ext=="pdf" else extract_text_from_docx(f) if ext=="docx" else None
                if not text:
                    return render(request,"core/voucher_upload.html",{"form":form,"error":"ملف غير مدعوم"})
                data = smart_parse(text)
                confirm_form = VoucherConfirmForm(initial={k:(v.strftime("%Y-%m-%d") if hasattr(v,"strftime") else v) for k,v in data.items()})
                return render(request,"core/voucher_confirm.html",{"confirm_form":confirm_form,"raw_text":text[:3000]})
            except Exception as e:
                return render(request,"core/voucher_upload.html",{"form":form,"error":f"خطأ أثناء القراءة: {e}"})
    else:
        form = VoucherUploadForm()
    return render(request,"core/voucher_upload.html",{"form":form})


@require_http_methods(["POST"])
def voucher_generate(request):
    form = VoucherConfirmForm(request.POST)
    if not form.is_valid():
        return render(request,"core/voucher_confirm.html",{"confirm_form":form,"error":"من فضلك راجع البيانات."})
    booking = form.cleaned_data
    today = datetime.today().date()
    html = render_to_string("core/voucher_template.html",{
        "booking": type("B",(),booking),"qr_data_url":"","today":today,
    })
    pdf_bytes = _render_pdf_from_template("core/voucher_template.html",{
        "booking": type("B",(),booking),"qr_data_url":"","today":today,
    })
    resp = HttpResponse(pdf_bytes,content_type="application/pdf")
    resp["Content-Disposition"]='inline; filename="voucher.pdf"'
    return resp
# ===================== Edit Payment =====================

@login_required
def hotel_payment_edit(request, booking_id):
    booking = get_object_or_404(HotelBooking, pk=booking_id)
    payment, created = Payment.objects.get_or_create(
        booking_hotel=booking,
        defaults={'net_price': booking.net or 0, 'sell_price': booking.sell or 0, 'paid_amount': 0}
    )

    if request.method == "POST":
        form = PaymentForm(request.POST, request.FILES, instance=payment)
        if form.is_valid():
            pay = form.save(commit=False)
            pay.booking_hotel = booking
            pay.full_clean()
            pay.save()

            # تحديث بيانات الحجز من الدفع
            booking.net, booking.sell = pay.net_price, pay.sell_price
            booking.save(update_fields=['net', 'sell'])

            return redirect('card_detail', booking.card_id)
    else:
        initial = {'net_price': booking.net or 0, 'sell_price': booking.sell or 0} if created else {}
        form = PaymentForm(instance=payment, initial=initial)

    return render(request, 'core/payment_form.html', {
        'card': booking.card,
        'booking': booking,
        'form': form
    })


# ===================== Dashboard Overview =====================

@login_required
def dashboard_overview(request):
    q_from     = parse_date(request.GET.get('from') or '')
    q_to       = parse_date(request.GET.get('to') or '')
    q_employee = (request.GET.get('employee') or '').strip()
    q_kind     = (request.GET.get('kind') or '').strip()   # hotel / flight / transfer / visa

    # صلاحيات
    card_filter, booking_filter = {}, {}
    if not request.user.is_superuser:
        card_filter['created_by'] = request.user
        booking_filter['card__created_by'] = request.user

    if q_from: booking_filter['created_at__date__gte'] = q_from
    if q_to:   booking_filter['created_at__date__lte'] = q_to
    if q_employee: booking_filter['employee_name__icontains'] = q_employee

    # Querysets
    hotel_qs_all    = HotelBooking.objects.filter(**booking_filter)
    flight_qs_all   = FlightBooking.objects.filter(**booking_filter)
    transfer_qs_all = TransferBooking.objects.filter(**booking_filter)
    visa_qs_all     = VisaBooking.objects.filter(**booking_filter)

    if q_kind == 'hotel':
        hotel_qs, flight_qs, transfer_qs, visa_qs = hotel_qs_all, HotelBooking.objects.none(), TransferBooking.objects.none(), VisaBooking.objects.none()
    elif q_kind == 'flight':
        hotel_qs, flight_qs, transfer_qs, visa_qs = HotelBooking.objects.none(), flight_qs_all, TransferBooking.objects.none(), VisaBooking.objects.none()
    elif q_kind == 'transfer':
        hotel_qs, flight_qs, transfer_qs, visa_qs = HotelBooking.objects.none(), FlightBooking.objects.none(), transfer_qs_all, VisaBooking.objects.none()
    elif q_kind == 'visa':
        hotel_qs, flight_qs, transfer_qs, visa_qs = HotelBooking.objects.none(), FlightBooking.objects.none(), TransferBooking.objects.none(), visa_qs_all
    else:
        hotel_qs, flight_qs, transfer_qs, visa_qs = hotel_qs_all, flight_qs_all, transfer_qs_all, visa_qs_all

    # KPIs
    total_cards = UniBookingCard.objects.filter(**card_filter).count()
    counts = {
        'hotels': hotel_qs_all.count(),
        'flights': flight_qs_all.count(),
        'transfers': transfer_qs_all.count(),
        'visas': visa_qs_all.count(),
        'all': hotel_qs_all.count() + flight_qs_all.count() + transfer_qs_all.count() + visa_qs_all.count(),
    }

    # إجماليات مالية (فنادق فقط)
    dec0 = V(Decimal('0.00'), output_field=DecimalField(max_digits=12, decimal_places=2))
    if q_kind not in ('', 'hotel'):
        total_net = total_sell = total_paid = total_remaining = Decimal('0.00')
        pays = Payment.objects.none()
    else:
        hotel_ids_for_pay = list(hotel_qs.values_list('id', flat=True))
        pays = Payment.objects.filter(booking_hotel_id__in=hotel_ids_for_pay)
        if q_from: pays = pays.filter(created_at__date__gte=q_from)
        if q_to:   pays = pays.filter(created_at__date__lte=q_to)

        totals = pays.aggregate(
            total_net=Coalesce(Sum('net_price'), dec0),
            total_sell=Coalesce(Sum('sell_price'), dec0),
            total_paid=Coalesce(Sum('paid_amount'), dec0),
        )
        total_net = totals['total_net'] or Decimal('0.00')
        total_sell = totals['total_sell'] or Decimal('0.00')
        total_paid = totals['total_paid'] or Decimal('0.00')
        total_remaining = total_sell - total_paid

    # المبيعات الشهرية
    months_map = {
        1:"يناير",2:"فبراير",3:"مارس",4:"إبريل",5:"مايو",6:"يونيو",
        7:"يوليو",8:"أغسطس",9:"سبتمبر",10:"أكتوبر",11:"نوفمبر",12:"ديسمبر"
    }
    monthly_raw = (
        pays.annotate(month=ExtractMonth('created_at'))
            .values('month')
            .annotate(sell=Coalesce(Sum('sell_price'), dec0),
                      net=Coalesce(Sum('net_price'), dec0))
            .order_by('month')
    )
    months, sales_data, net_data = [], [], []
    for r in monthly_raw:
        m = r['month']
        if not m: continue
        months.append(months_map.get(m, str(m)))
        sales_data.append(float(r['sell'] or 0))
        net_data.append(float(r['net'] or 0))

    # أحدث 5 حجوزات
    latest = sorted(
        list(hotel_qs) + list(flight_qs) + list(transfer_qs) + list(visa_qs),
        key=lambda b: b.created_at, reverse=True
    )[:5]
    for b in latest:
        setattr(b, 'display_kind', b._meta.model_name)

    # Top Agents & Customers
    agent_counter, customer_counter = Counter(), Counter()
    def _count_from_qs(qs):
        for r in qs.values('employee_name','card__customer_name'):
            if r['employee_name']: agent_counter[r['employee_name']] += 1
            if r['card__customer_name']: customer_counter[r['card__customer_name']] += 1
    _count_from_qs(hotel_qs); _count_from_qs(flight_qs); _count_from_qs(transfer_qs); _count_from_qs(visa_qs)
    top_agents = agent_counter.most_common(5)
    top_customers = customer_counter.most_common(5)

    # توزيع أنواع الحجوزات
    booking_types_data = [hotel_qs.count(), flight_qs.count(), transfer_qs.count(), visa_qs.count()]

    # قائمة الموظفين
    employees_set = set(
        list(hotel_qs_all.values_list('employee_name', flat=True)) +
        list(flight_qs_all.values_list('employee_name', flat=True)) +
        list(transfer_qs_all.values_list('employee_name', flat=True)) +
        list(visa_qs_all.values_list('employee_name', flat=True))
    )
    employees = sorted([e for e in employees_set if (e or '').strip()])

    return render(request,'core/dashboard_overview.html',{
        'total_cards': total_cards,
        'counts': counts,
        'total_net': total_net, 'total_sell': total_sell,
        'total_paid': total_paid, 'total_remaining': total_remaining,
        'months': months, 'sales_data': sales_data, 'net_data': net_data,
        'booking_types_data': booking_types_data,
        'latest': latest, 'top_agents': top_agents, 'top_customers': top_customers,
        'q_from': request.GET.get('from',''), 'q_to': request.GET.get('to',''),
        'q_employee': q_employee, 'q_kind': q_kind, 'employees': employees,
    })
# في نهاية ملف core/views.py

from django.views.decorators.http import require_POST

@login_required
def payment_edit(request, payment_pk):
    payment = get_object_or_404(Payment, pk=payment_pk)
    booking = payment.booking_hotel

    # شروط الأمان
    if not request.user.is_superuser:
        messages.error(request, "ليس لديك صلاحية لتعديل الدفعات.")
        return redirect('card_detail', pk=booking.card_id)
    
    if not payment.is_editable:
        messages.error(request, "لا يمكن تعديل هذه الدفعة بعد مرور 24 ساعة على إنشائها.")
        return redirect('card_detail', pk=booking.card_id)

    if request.method == 'POST':
        # عند التعديل، لا نمرر معلومات عن الدفعات السابقة لأن الفورم لا يحتاجها
        form = PaymentForm(request.POST, request.FILES, instance=payment, booking=booking)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تعديل الدفعة بنجاح.")
            return redirect('card_detail', pk=booking.card_id)
    else:
        form = PaymentForm(instance=payment, booking=booking)

    return render(request, 'core/payment_edit_form.html', {'form': form, 'payment': payment})


@login_required
@require_POST # لضمان أن الحذف يتم فقط عبر طلب POST
def payment_delete(request, payment_pk):
    payment = get_object_or_404(Payment, pk=payment_pk)
    card_id = payment.booking_hotel.card_id

    # شروط الأمان
    if not request.user.is_superuser:
        messages.error(request, "ليس لديك صلاحية لحذف الدفعات.")
        return redirect('card_detail', pk=card_id)

    if not payment.is_editable:
        messages.error(request, "لا يمكن حذف هذه الدفعة بعد مرور 24 ساعة على إنشائها.")
        return redirect('card_detail', pk=card_id)

    payment.delete()
    messages.success(request, "تم حذف الدفعة بنجاح.")
    return redirect('card_detail', pk=card_id)